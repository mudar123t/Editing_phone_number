import openpyxl
t = openpyxl.load_workbook("UAE.xlsx")

sh = t.sheetnames
sh1 = t['Sheet2']
row = sh1.max_row
col = sh1.max_column
nrows = 0
for i in range(1, row+1):
    i = sh1.cell(i, 1).value
    i = str(i)

    if i.startswith('5'):
        i = '971' + i
        b8 = "".join(i)
        i = b8

    elif i.startswith('971 0'):
        l9 = list(i)
        l9.remove(" ")
        l9.remove("0")
        b9 = "".join(l9)
        i = b9

    elif i.startswith('971 '):
        l10 = list(i)
        l10.remove(" ")
        b10 = "".join(l10)
        i = b10

    elif i.startswith('0971'):
        l6 = list(i)
        l6.remove("0")
        b6 = "".join(l6)
        i = b6

    elif i.startswith('00971'):
        l7 = list(i)
        l7.remove("0")
        l7.remove("0")
        b7 = "".join(l7)
        i = b7

    elif i.startswith('9710'):
        l5 = list(i)
        l5.remove("0")
        b5 = "".join(l5)
        i = b5

    elif i.startswith('+'):
        l = list(i)
        l.remove('+')
        b = "".join(l)
        i = b

    elif i.startswith('05'):
        t = '971' + i
        l1 = list(t)
        l1.remove("0")
        b1 = "".join(l1)
        i = b1

    elif i.startswith('971-971'):
        l2 = list(i)
        p1 = 0
        for nu in l2:
            l2.pop(0)
            p1 = p1+1
            if p1 == 4:
                break
        b2 = "".join(l2)
        i = b2

    elif i.startswith('971-5'):
        l3 = list(i)
        l3.remove('-')
        b3 = "".join(l3)
        i = b3

    elif i.startswith('971-0'):
        l4 = list(i)
        l4.remove('-')
        l4.remove('0')
        b4 = "".join(l4)
        i = b4

    nrows = nrows + 1
    print(i)

print(nrows)