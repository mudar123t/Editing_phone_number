import openpyxl

t = openpyxl.load_workbook("oman.xlsx")

sh = t.sheetnames
sh1 = t['sheet1']
row = sh1.max_row
col = sh1.max_column
nrows = 0
for i in range(1, row + 1):
    i = sh1.cell(i, 5).value
    i = str(i)

    if i.startswith('7'):
        i = '968' + i

    elif i.startswith("91"):
        i = '968' + i

    elif i.startswith("92"):
        i = '968' + i

    elif i.startswith("93"):
        i = '968' + i

    elif i.startswith("94"):
        i = '968' + i

    elif i.startswith("95"):
        i = '968' + i

    elif i.startswith("97"):
        i = '968' + i

    elif i.startswith("98"):
        i = '968' + i

    elif i.startswith("99"):
        i = '968' + i

    elif i.startswith('0968'):
        l6 = list(i)
        l6.remove("0")
        b6 = "".join(l6)
        i = b6

    elif i.startswith('00968'):
        l7 = list(i)
        l7.remove("0")
        l7.remove("0")
        b7 = "".join(l7)
        i = b7

    elif i.startswith('9680'):
        l5 = list(i)
        l5.remove("0")
        b5 = "".join(l5)
        i = b5

    elif i.startswith('+'):
        l = list(i)
        l.remove('+')
        b = "".join(l)
        i = b

    elif i.startswith('09'):
        t = '968' + i
        l1 = list(t)
        l1.remove("0")
        b1 = "".join(l1)
        i = b1

    elif i.startswith('968-968'):
        l2 = list(i)
        p1 = 0
        for nu in l2:
            l2.pop(0)
            p1 = p1 + 1
            if p1 == 4:
                break
        b2 = "".join(l2)
        i = b2

    elif i.startswith('968-9'):
        l3 = list(i)
        l3.remove('-')
        b3 = "".join(l3)
        i = b3

    elif i.startswith('968-0'):
        l4 = list(i)
        l4.remove('-')
        l4.remove('0')
        b4 = "".join(l4)
        i = b4

    nrows = nrows + 1
    print(i)

print(nrows)
